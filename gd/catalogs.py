"""Catalog loading utilities (hoja Datos)."""
from __future__ import annotations

from typing import Dict, Tuple

from openpyxl.utils import column_index_from_string, get_column_letter

from .config import COLS
from .excel import (
    get_unique_list_from_column,
    load_workbook,
    get_ws_datos,
    find_column_by_header,
)
from .models import Catalogs


def load_dependency_mapping(wb=None) -> Dict[str, str]:
    """Return mapping: célula -> header de descripción."""
    wb = wb or load_workbook()
    ws_d = get_ws_datos(wb)
    col_cel_dep_idx = find_column_by_header(ws_d, "Celula Dependencia", header_row=1)
    col_desc_dep_idx = find_column_by_header(ws_d, "Celula Descripcion Dependencia", header_row=1)
    mapping: Dict[str, str] = {}
    if not col_cel_dep_idx or not col_desc_dep_idx:
        return mapping

    for row in range(2, ws_d.max_row + 1):
        cel_name = ws_d.cell(row=row, column=col_cel_dep_idx).value
        desc_header = ws_d.cell(row=row, column=col_desc_dep_idx).value
        if cel_name and desc_header:
            mapping[str(cel_name).strip()] = str(desc_header).strip()
    return mapping


def _load_celula_tren_map(ws_d) -> Dict[str, str]:
    celula_tren_map: Dict[str, str] = {}
    col_tren_idx = column_index_from_string("E")
    col_cel_idx = column_index_from_string("F")
    for row in range(2, ws_d.max_row + 1):
        tren_val = ws_d.cell(row=row, column=col_tren_idx).value
        cel_val = ws_d.cell(row=row, column=col_cel_idx).value
        if tren_val and cel_val:
            celula_tren_map[str(cel_val).strip()] = str(tren_val).strip()
    return celula_tren_map


def load_catalogs() -> Catalogs:
    """Load dropdown catalog values and dependency mappings from hoja Datos."""
    wb = load_workbook()
    ws_d = get_ws_datos(wb)

    estados_list = get_unique_list_from_column(ws_d, "A")
    priorizacion_list = get_unique_list_from_column(ws_d, "B")
    responsables_list = get_unique_list_from_column(ws_d, "C")
    areas_list = get_unique_list_from_column(ws_d, "D")

    celula_tren_map = _load_celula_tren_map(ws_d)
    area_tren_coe_list = sorted({v for v in celula_tren_map.values()})
    celulas_dep_list = sorted({k for k in celula_tren_map.keys()})

    iniciativas_list = []
    col_ini_idx = find_column_by_header(ws_d, "Iniciativa Estrategica", header_row=1)
    if col_ini_idx:
        col_ini_letter = get_column_letter(col_ini_idx)
        iniciativas_list = get_unique_list_from_column(ws_d, col_ini_letter)

    dep_mapping = load_dependency_mapping(wb)
    if dep_mapping:
        celulas_dep_from_mapping = sorted(dep_mapping.keys())
        celulas_dep_list = sorted(set(celulas_dep_list) | set(celulas_dep_from_mapping))

    return Catalogs(
        estados=estados_list,
        q_rad=priorizacion_list,
        responsables=responsables_list,
        areas=areas_list,
        area_tren_coe=area_tren_coe_list,
        celulas_dep=celulas_dep_list,
        iniciativas=iniciativas_list,
        dependency_mapping=dep_mapping,
        celula_tren_map=celula_tren_map,
    )
