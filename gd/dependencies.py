"""Dependency helpers for applying and aggregating flags/descriptions."""
from __future__ import annotations

from typing import Iterable, List, Sequence

from openpyxl.utils import column_index_from_string

from .config import (
    CARD_BORDER,
    DARK_COLOR,
    DESC_END_COL,
    DESC_START_COL,
    FLAG_END_COL,
    FLAG_START_COL,
    COLS,
)
from .excel import find_column_by_header_in_range, get_header_row_proyectos
from .models import Dependency


def compute_dep_aggregates(dep_list: Sequence[Dependency]):
    flags = [
        (dep.codigo or "").strip().upper()
        for dep in dep_list
        if (dep.equipo or "").strip()
    ]
    flags = [f for f in flags if f in ("L", "P")]
    total_dep = len(flags)
    total_L = sum(1 for f in flags if f == "L")
    total_P = sum(1 for f in flags if f == "P")
    cubrimiento = (total_P / total_dep) if total_dep else 0.0
    return total_dep, total_L, total_P, cubrimiento


def write_dep_aggregates(ws, row: int, dep_list: Sequence[Dependency]):
    total_dep, total_L, total_P, cub = compute_dep_aggregates(dep_list)
    cn = column_index_from_string(COLS["TOTAL_DEP"])
    co = column_index_from_string(COLS["TOTAL_L"])
    cp = column_index_from_string(COLS["TOTAL_P"])
    cq = column_index_from_string(COLS["CUBRIMIENTO_DEP"])

    ws.cell(row=row, column=cn).value = total_dep
    ws.cell(row=row, column=co).value = total_L
    ws.cell(row=row, column=cp).value = total_P
    ws.cell(row=row, column=cq).value = cub


def apply_dependencies_to_row(ws, row: int, dep_list: Sequence[Dependency], dep_mapping: dict):
    header_row = get_header_row_proyectos(ws)

    for equipo, desc_header in dep_mapping.items():
        flag_col_idx = find_column_by_header_in_range(
            ws, equipo, FLAG_START_COL, FLAG_END_COL, header_row
        )
        if flag_col_idx:
            ws.cell(row=row, column=flag_col_idx).value = None

        if desc_header:
            desc_col_idx = find_column_by_header_in_range(
                ws, desc_header, DESC_START_COL, DESC_END_COL, header_row
            )
            if desc_col_idx:
                ws.cell(row=row, column=desc_col_idx).value = None

    for dep in dep_list:
        equipo = (dep.equipo or "").strip()
        codigo = (dep.codigo or "").strip().upper()
        texto = (dep.descripcion or "").strip()

        if not equipo or codigo not in ("P", "L"):
            continue

        desc_header = dep_mapping.get(equipo)

        flag_col_idx = find_column_by_header_in_range(
            ws, equipo, FLAG_START_COL, FLAG_END_COL, header_row
        )
        if flag_col_idx:
            ws.cell(row=row, column=flag_col_idx).value = codigo

        if desc_header:
            desc_col_idx = find_column_by_header_in_range(
                ws, desc_header, DESC_START_COL, DESC_END_COL, header_row
            )
            if desc_col_idx and texto:
                ws.cell(row=row, column=desc_col_idx).value = texto

    write_dep_aggregates(ws, row, dep_list)


def dep_semaforo(total_dep: int, total_L: int, total_P: int):
    if total_dep == 0:
        return "#bdc3c7", "Sin dependencias registradas"
    if total_P == 0 and total_L > 0:
        return "#2ecc71", "Todas las dependencias negociadas (L)"
    if total_L == 0 and total_P > 0:
        return "#e74c3c", "Todas las dependencias pendientes (P)"
    return "#f1c40f", "Mix de dependencias negociadas (L) y pendientes (P)"


def build_semaforo_block(total_dep: int, total_L: int, total_P: int, title: str = "Dependencias"):
    _color, text = dep_semaforo(total_dep, total_L, total_P)

    if total_dep == 0:
        state = "apagado"
    elif total_P == 0 and total_L > 0:
        state = "verde"
    elif total_L == 0 and total_P > 0:
        state = "rojo"
    else:
        state = "amarillo"

    def light(color_hex: str, active: bool) -> str:
        fill = color_hex if active else "#e0e0e0"
        return f"""
        <div style="
            width:22px;
            height:22px;
            border-radius:50%;
            background-color:{fill};
            margin:4px auto;
            border:1px solid #999;
        "></div>
        """

    red_light = light("#e74c3c", state == "rojo")
    yellow_light = light("#f1c40f", state == "amarillo")
    green_light = light("#2ecc71", state == "verde")

    html = f"""
    <div style="
        font-family:Segoe UI, Arial;
        font-size:13px;
        color:{DARK_COLOR};
        text-align:center;
    ">
      <div style="font-weight:600; margin-bottom:8px;">{title}</div>
      <div style="
          width:60px;
          margin:0 auto 8px auto;
          padding:8px 0;
          border-radius:30px;
          background-color:white;
          border:1px solid {CARD_BORDER};
      ">
        {red_light}
        {yellow_light}
        {green_light}
      </div>
      <div style="font-size:12px; line-height:1.4; max-width:220px; margin:0 auto;">
        {text}
      </div>
    </div>
    """
    return html
