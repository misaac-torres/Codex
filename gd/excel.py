"""Excel helpers extracted from the original notebook."""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, Tuple

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from .config import (
    COLS,
    EXCEL_PATH,
    HEADER_ROW_PROYECTOS,
    SHEET_DATOS,
    SHEET_PROYECTOS,
    SHEET_SUG,
    START_ROW_PROYECTOS,
    ensure_required_sheets,
)


# ---------------------------------------------------------------------------
# Workbook access
# ---------------------------------------------------------------------------

def load_workbook(path: Path = EXCEL_PATH) -> openpyxl.Workbook:
    """Open the workbook and validate required sheets."""
    if not path.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {path}")
    # Guard against unsupported binary formats (e.g., .xlsb) early so the error
    # message is clearer for Swagger/CLI users.
    allowed_suffixes = {".xlsx", ".xlsm", ".xltx", ".xltm"}
    if path.suffix.lower() not in allowed_suffixes:
        raise ValueError(
            "Formato de Excel no soportado. Usa un archivo .xlsx/.xlsm (no binario como .xlsb): "
            f"{path}"
        )

    try:
        wb = openpyxl.load_workbook(path, keep_vba=False)
    except openpyxl.utils.exceptions.InvalidFileException as exc:
        raise ValueError(
            "No se pudo abrir el Excel. Asegúrate de que no sea un archivo binario (.xlsb) y de que esté válido: "
            f"{path}"
        ) from exc
    ensure_required_sheets(wb.sheetnames)
    return wb


def get_ws_proyectos(wb: openpyxl.Workbook | None = None):
    wb = wb or load_workbook()
    return wb[SHEET_PROYECTOS]


def get_ws_datos(wb: openpyxl.Workbook | None = None):
    wb = wb or load_workbook()
    return wb[SHEET_DATOS]


def get_ws_sugerencias(wb: openpyxl.Workbook | None = None):
    """Return the suggestion sheet, creating it if needed."""
    wb = wb or load_workbook()
    if SHEET_SUG not in wb.sheetnames:
        ws = wb.create_sheet(SHEET_SUG)
        ws["A1"] = "Usuario"
        ws["B1"] = "Sugerencia"
    else:
        ws = wb[SHEET_SUG]
        if ws.max_row == 1 and ws["A1"].value is None:
            ws["A1"] = "Usuario"
            ws["B1"] = "Sugerencia"
    return ws


# ---------------------------------------------------------------------------
# Column and value helpers
# ---------------------------------------------------------------------------

def get_header_row_proyectos(ws) -> int:
    """Detect header row by locating the ID column header."""
    id_col_idx = column_index_from_string(COLS["ID"])
    for r in range(1, START_ROW_PROYECTOS):
        v = ws.cell(row=r, column=id_col_idx).value
        if isinstance(v, str) and v.strip().lower() == "id":
            return r
    return HEADER_ROW_PROYECTOS


def get_unique_list_from_column(ws, col_letter: str, start_row: int = 2):
    values = set()
    col_idx = column_index_from_string(col_letter)
    for row in range(start_row, ws.max_row + 1):
        value = ws.cell(row=row, column=col_idx).value
        if value not in (None, ""):
            values.add(str(value))
    return sorted(values)


def get_next_row_and_id(ws, id_col_letter: str = "A", start_row: int = START_ROW_PROYECTOS) -> Tuple[int, int]:
    id_col_idx = column_index_from_string(id_col_letter)
    max_row_used = 0
    max_id_found = 0

    for row in range(start_row, ws.max_row + 1):
        val = ws.cell(row=row, column=id_col_idx).value
        if val not in (None, ""):
            max_row_used = row
            if isinstance(val, (int, float)):
                max_id_found = max(max_id_found, int(val))

    next_row = start_row if max_row_used == 0 else max_row_used + 1
    next_id = max_id_found + 1 if max_id_found > 0 else 1
    return next_row, next_id


def find_column_by_header(ws, header_name: str, header_row: int = 1):
    if not header_name:
        return None
    target = str(header_name).strip().lower()
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is None:
            continue
        if str(val).strip().lower() == target:
            return col_idx
    return None


def find_column_by_header_in_range(ws, header_name: str, start_col_idx: int, end_col_idx: int, header_row: int):
    if not header_name:
        return None
    target = str(header_name).strip().lower()
    for col_idx in range(start_col_idx, end_col_idx + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val is None:
            continue
        if str(val).strip().lower() == target:
            return col_idx
    return None


def find_area_tren_coe_col(ws):
    header_row = get_header_row_proyectos(ws)
    candidate_idx = None
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if not val:
            continue
        s = str(val).strip().lower()
        if "tren" in s and "coe" in s:
            return col_idx
        if s in ("area tren coe", "area/tren/coe"):
            candidate_idx = col_idx
    return candidate_idx


def to_num_cell(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "":
        return 0.0
    s = s.replace("%", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return 0.0


def iter_rows(ws, start_row: int, end_row: int, col_letters: Iterable[str]):
    col_indices = [column_index_from_string(c) for c in col_letters]
    for row in range(start_row, end_row + 1):
        yield [ws.cell(row=row, column=idx).value for idx in col_indices]


def column_letter(col_idx: int) -> str:
    return get_column_letter(col_idx)
