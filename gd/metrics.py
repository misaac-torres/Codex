"""Aggregate metrics used by the original Métricas tab."""
from __future__ import annotations

from openpyxl.utils import column_index_from_string

from .config import COLS, EXCEL_PATH, START_ROW_PROYECTOS
from .excel import (
    find_column_by_header_in_range,
    get_header_row_proyectos,
    get_ws_proyectos,
    load_workbook,
    to_num_cell,
)


def compute_metrics(scope: str = "all", filter_value: str | None = None, dep_mapping: dict | None = None, celula_tren_map: dict | None = None, path=EXCEL_PATH):
    wb = load_workbook(path)
    ws = get_ws_proyectos(wb)
    header_row = get_header_row_proyectos(ws)

    name_col_idx = column_index_from_string(COLS["NOMBRE_PROYECTO"])
    prior_col_idx = column_index_from_string(COLS["PRIORIZADO"])
    av_col_idx = column_index_from_string(COLS["AVANCE"])
    cn_idx = column_index_from_string(COLS["TOTAL_DEP"])
    co_idx = column_index_from_string(COLS["TOTAL_L"])
    cp_idx = column_index_from_string(COLS["TOTAL_P"])

    cel_flag_col_idx = None
    if scope == "celula" and filter_value:
        cel_flag_col_idx = find_column_by_header_in_range(
            ws, filter_value, column_index_from_string("R"), column_index_from_string("BB"), header_row
        )

    total_projects = 0
    total_dep = 0.0
    total_L = 0.0
    total_P = 0.0
    sum_avance = 0.0
    pri_count = 0
    pri_avance_sum = 0.0
    no_pri_count = 0
    no_pri_avance_sum = 0.0

    for row in range(START_ROW_PROYECTOS, ws.max_row + 1):
        name = ws.cell(row=row, column=name_col_idx).value
        if not name:
            continue

        if scope == "area" and filter_value:
            tren_match = False
            dep_mapping = dep_mapping or {}
            celula_tren_map = celula_tren_map or {}
            for equipo in dep_mapping.keys():
                flag_col_idx = find_column_by_header_in_range(
                    ws, equipo, column_index_from_string("R"), column_index_from_string("BB"), header_row
                )
                if not flag_col_idx:
                    continue
                flag_val = ws.cell(row=row, column=flag_col_idx).value
                if not flag_val:
                    continue
                flag_up = str(flag_val).strip().upper()
                if flag_up not in ("P", "L"):
                    continue
                tren_val = celula_tren_map.get(equipo)
                if tren_val and str(tren_val).strip() == str(filter_value).strip():
                    tren_match = True
                    break
            if not tren_match:
                continue

        if scope == "celula" and filter_value:
            if not cel_flag_col_idx:
                continue
            flag_val = ws.cell(row=row, column=cel_flag_col_idx).value
            if not flag_val or str(flag_val).strip().upper() not in ("P", "L"):
                continue

        total_projects += 1

        dep_val = to_num_cell(ws.cell(row=row, column=cn_idx).value)
        L_val = to_num_cell(ws.cell(row=row, column=co_idx).value)
        P_val = to_num_cell(ws.cell(row=row, column=cp_idx).value)

        total_dep += dep_val
        total_L += L_val
        total_P += P_val

        av_val = to_num_cell(ws.cell(row=row, column=av_col_idx).value)
        sum_avance += av_val

        pri_val = ws.cell(row=row, column=prior_col_idx).value
        pri_str = str(pri_val).strip().upper() if pri_val not in (None, "") else ""

        if pri_str == "SI":
            pri_count += 1
            pri_avance_sum += av_val
        else:
            no_pri_count += 1
            no_pri_avance_sum += av_val

    avg_avance = (sum_avance / total_projects) if total_projects > 0 else 0.0
    avg_pri = (pri_avance_sum / pri_count) if pri_count > 0 else 0.0
    avg_no_pri = (no_pri_avance_sum / no_pri_count) if no_pri_count > 0 else 0.0
    cobertura_pct = (total_P / total_dep * 100.0) if total_dep > 0 else 0.0

    return {
        "total_projects": int(total_projects),
        "total_dep": float(total_dep),
        "total_L": float(total_L),
        "total_P": float(total_P),
        "cobertura_pct": float(cobertura_pct),
        "avg_avance": float(avg_avance),
        "avg_pri": float(avg_pri),
        "avg_no_pri": float(avg_no_pri),
        "num_pri": int(pri_count),
    }
