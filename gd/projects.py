"""Project CRUD helpers based on the Excel workbook."""
from __future__ import annotations

from typing import Iterable, Sequence

from openpyxl.utils import column_index_from_string

from .config import COLS, EXCEL_PATH, START_ROW_PROYECTOS
from .dependencies import apply_dependencies_to_row
from .excel import (
    find_column_by_header_in_range,
    get_header_row_proyectos,
    get_next_row_and_id,
    get_ws_proyectos,
    load_workbook,
    to_num_cell,
)
from .models import Dependency, Project


def write_project_with_dependencies(project: Project, dep_list: Sequence[Dependency], dep_mapping: dict, path=EXCEL_PATH):
    """Insert a new project row and apply dependencies + aggregates."""
    wb = load_workbook(path)
    ws = get_ws_proyectos(wb)

    next_row, next_id = get_next_row_and_id(
        ws, id_col_letter=COLS["ID"], start_row=START_ROW_PROYECTOS
    )

    row_data = project.to_row_mapping()
    row_data["ID"] = next_id

    for field, col_letter in COLS.items():
        if field in row_data:
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=next_row, column=col_idx).value = row_data.get(field)

    apply_dependencies_to_row(ws, next_row, dep_list, dep_mapping)

    wb.save(path)
    return next_row, next_id


def get_all_project_names(path=EXCEL_PATH):
    wb = load_workbook(path)
    ws = get_ws_proyectos(wb)
    name_col_idx = column_index_from_string(COLS["NOMBRE_PROYECTO"])
    names = set()
    for row in range(START_ROW_PROYECTOS, ws.max_row + 1):
        val = ws.cell(row=row, column=name_col_idx).value
        if val not in (None, ""):
            names.add(str(val).strip())
    return sorted(names)


def summarize_by_equipo(equipo_name: str, dep_mapping: dict, path=EXCEL_PATH):
    wb = load_workbook(path)
    ws = get_ws_proyectos(wb)
    header_row = get_header_row_proyectos(ws)

    col_flag_idx = find_column_by_header_in_range(
        ws, equipo_name, column_index_from_string("R"), column_index_from_string("BB"), header_row
    )
    if not col_flag_idx:
        return {"found": False, "msg": f"No se encontró la columna '{equipo_name}' en R:BB."}

    name_col_idx = column_index_from_string(COLS["NOMBRE_PROYECTO"])
    q_col_idx = column_index_from_string(COLS["Q_RADICADO"])

    total = pendientes = negociadas = 0
    rows = []

    for row in range(START_ROW_PROYECTOS, ws.max_row + 1):
        flag = ws.cell(row=row, column=col_flag_idx).value
        if flag is None or str(flag).strip() == "":
            continue
        flag_up = str(flag).strip().upper()
        if flag_up not in ("P", "L"):
            continue

        total += 1
        if flag_up == "P":
            pendientes += 1
        else:
            negociadas += 1

        nombre = ws.cell(row=row, column=name_col_idx).value
        qrad = ws.cell(row=row, column=q_col_idx).value
        rows.append({"fila": row, "Q_RADICADO": qrad, "PROYECTO": nombre, "FLAG": flag_up})

    pct_pend = (pendientes / total * 100) if total > 0 else 0.0
    return {
        "found": True,
        "equipo": equipo_name,
        "total": total,
        "pendientes": pendientes,
        "negociadas": negociadas,
        "pct_pendientes": pct_pend,
        "rows": rows,
    }


def summarize_by_proyecto(nombre_proyecto: str, dep_mapping: dict, path=EXCEL_PATH):
    wb = load_workbook(path)
    ws = get_ws_proyectos(wb)
    header_row = get_header_row_proyectos(ws)

    name_col_idx = column_index_from_string(COLS["NOMBRE_PROYECTO"])
    q_col_idx = column_index_from_string(COLS["Q_RADICADO"])

    target_row = None
    for row in range(START_ROW_PROYECTOS, ws.max_row + 1):
        val = ws.cell(row=row, column=name_col_idx).value
        if val and str(val).strip() == nombre_proyecto:
            target_row = row
            break

    if not target_row:
        return {"found": False, "msg": f"No se encontró el proyecto '{nombre_proyecto}'."}

    detalles = []

    for equipo, desc_header in dep_mapping.items():
        flag_col_idx = find_column_by_header_in_range(
            ws, equipo, column_index_from_string("R"), column_index_from_string("BB"), header_row
        )
        if not flag_col_idx:
            continue

        flag = ws.cell(row=target_row, column=flag_col_idx).value
        if flag is None or str(flag).strip() == "":
            continue

        flag_up = str(flag).strip().upper()
        if flag_up not in ("P", "L"):
            continue

        desc = ""
        if desc_header:
            desc_col_idx = find_column_by_header_in_range(
                ws, desc_header, column_index_from_string("BC"), column_index_from_string("CM"), header_row
            )
            if desc_col_idx:
                desc = ws.cell(row=target_row, column=desc_col_idx).value or ""

        detalles.append({"equipo": equipo, "FLAG": flag_up, "descripcion": desc})

    total = len(detalles)
    pendientes = sum(1 for d in detalles if d["FLAG"] == "P")
    negociadas = sum(1 for d in detalles if d["FLAG"] == "L")
    pct_pend = (pendientes / total * 100) if total > 0 else 0.0

    lb_col = column_index_from_string(COLS["LINEA_BASE"])
    av_col = column_index_from_string(COLS["AVANCE"])
    est_col = column_index_from_string(COLS["ESTIMADO_AVANCE"])

    lb = to_num_cell(ws.cell(row=target_row, column=lb_col).value)
    av = to_num_cell(ws.cell(row=target_row, column=av_col).value)
    est = to_num_cell(ws.cell(row=target_row, column=est_col).value)
    qrad = ws.cell(row=target_row, column=q_col_idx).value

    cn = column_index_from_string(COLS["TOTAL_DEP"])
    co = column_index_from_string(COLS["TOTAL_L"])
    cp = column_index_from_string(COLS["TOTAL_P"])
    cq = column_index_from_string(COLS["CUBRIMIENTO_DEP"])

    total_dep_xl = to_num_cell(ws.cell(row=target_row, column=cn).value)
    total_L_xl = to_num_cell(ws.cell(row=target_row, column=co).value)
    total_P_xl = to_num_cell(ws.cell(row=target_row, column=cp).value)
    cub_xl = to_num_cell(ws.cell(row=target_row, column=cq).value)

    return {
        "found": True,
        "fila": target_row,
        "proyecto": nombre_proyecto,
        "Q_RADICADO": qrad,
        "total_dep": total,
        "pendientes": pendientes,
        "negociadas": negociadas,
        "pct_pendientes": pct_pend,
        "detalles": detalles,
        "linea_base": float(lb),
        "avance": float(av),
        "estimado": float(est),
        "total_dep_xl": total_dep_xl,
        "total_L_xl": total_L_xl,
        "total_P_xl": total_P_xl,
        "cub_xl": float(cub_xl),
    }


def update_project_row_and_dependencies(
    row: int,
    avance: float | None,
    estimado: float | None,
    dep_list: Sequence[Dependency],
    dep_mapping: dict,
    path=EXCEL_PATH,
):
    wb = load_workbook(path)
    ws = get_ws_proyectos(wb)

    lb_col = column_index_from_string(COLS["LINEA_BASE"])
    av_col = column_index_from_string(COLS["AVANCE"])
    est_col = column_index_from_string(COLS["ESTIMADO_AVANCE"])
    pct_col = column_index_from_string(COLS["PORC_CUMPLIMIENTO"])

    linea_base = to_num_cell(ws.cell(row=row, column=lb_col).value)
    old_av = to_num_cell(ws.cell(row=row, column=av_col).value)
    old_es = to_num_cell(ws.cell(row=row, column=est_col).value)

    new_av = float(avance) if avance is not None else float(old_av)
    new_es = float(estimado) if estimado is not None else float(old_es)

    ws.cell(row=row, column=av_col).value = new_av
    ws.cell(row=row, column=est_col).value = new_es

    pct_cumpl = (new_av / new_es) if new_es > 0 else 0.0
    ws.cell(row=row, column=pct_col).value = pct_cumpl

    apply_dependencies_to_row(ws, row, dep_list, dep_mapping)

    wb.save(path)

    var_vs_lb = new_av - float(linea_base)
    return {
        "linea_base": float(linea_base),
        "avance": new_av,
        "estimado": new_es,
        "pct_cumpl": pct_cumpl * 100,
        "var_vs_lb_pp": var_vs_lb * 100,
    }
